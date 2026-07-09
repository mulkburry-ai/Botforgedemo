# ============================================================
# BOTFORGE — scheduler.py
# Orqa fon vazifalari (APScheduler)
# 1. Har daqiqa  → vaqtli chegirmalarni tekshirish
# 2. Har kuni 09:00 → kam stokli mahsulotlar haqida ogohlantirish
# 3. Har kuni 03:00 → avtomatik Excel bekap
# ============================================================

import logging
import io
from datetime import datetime
from aiogram import Bot
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger

from config import ADMIN_ID
from db import (
    get_expired_discounts, deactivate_discount,
    get_low_stock_products, get_setting,
    get_all_orders_for_export, get_all_products_for_export,
)
from utils import fmt_date_short

logger = logging.getLogger(__name__)


# ============================================================
# VAZIFA 1 — Vaqtli chegirmalarni tekshirish
# Har 1 daqiqada ishlaydi
# ============================================================

async def check_expired_discounts() -> None:
    """
    Vaqti o'tgan chegirmalarni o'chiradi.
    Misol: admin 24 soatlik chegirma qo'ydi →
    24 soat o'tgandan keyin bu funksiya uni o'chiradi.
    """
    expired_ids = await get_expired_discounts()
    if not expired_ids:
        return

    for product_id in expired_ids:
        await deactivate_discount(product_id)
        logger.info(f"Chegirma o'chirildi: mahsulot #{product_id}")

    logger.info(f"Jami {len(expired_ids)} ta chegirma o'chirildi")


# ============================================================
# VAZIFA 2 — Kam stok ogohlantirish
# Har kuni soat 09:00 da ishlaydi
# ============================================================

async def low_stock_alert(bot: Bot) -> None:
    """
    Stok kam mahsulotlar haqida adminga xabar yuboradi.
    Chegara: settings jadvalidagi low_stock_threshold qiymati.
    """
    threshold = int(await get_setting("low_stock_threshold", "5"))
    products  = await get_low_stock_products(threshold)

    if not products:
        return

    lines = [f"⚠️ *Stok ogohlantirishi*\n\nQuyidagi mahsulotlar {threshold} dona va undan kam:\n"]
    for p in products:
        stock_icon = "🚫" if p["stock_qty"] == 0 else "⚠️"
        lines.append(f"{stock_icon} `#{p['id']}` {p['name']} — *{p['stock_qty']} dona*")

    try:
        await bot.send_message(
            chat_id=ADMIN_ID,
            text="\n".join(lines),
            parse_mode="Markdown"
        )
        logger.info(f"Stok ogohlantirishi yuborildi: {len(products)} ta mahsulot")
    except Exception as e:
        logger.error(f"Stok ogohlantirishi yuborishda xato: {e}")


# ============================================================
# VAZIFA 3 — Avtomatik Excel bekap
# Har kuni soat 03:00 da ishlaydi
# ============================================================

async def daily_backup(bot: Bot) -> None:
    """
    Buyurtmalar va mahsulotlarni Excel faylga eksport qiladi.
    Adminga yuboriladi.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from aiogram.types import BufferedInputFile

        today = fmt_date_short(datetime.now())

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E86AB")

        # ── Buyurtmalar ─────────────────────────────────────
        orders = await get_all_orders_for_export()
        wb1    = openpyxl.Workbook()
        ws1    = wb1.active
        ws1.title = "Buyurtmalar"

        headers1 = [
            "Kod", "Holat", "Summa", "Manzil",
            "Telefon", "Izoh", "Sana",
            "Ism", "Username", "Foydalanuvchi tel"
        ]
        ws1.append(headers1)
        for cell in ws1[1]:
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center")

        for o in orders:
            ws1.append([
                o["order_code"],
                o["status"],
                float(o["total_price"] or 0),
                o["address"]  or "",
                o["order_phone"] or "",
                o["note"]     or "",
                o["created_at"].strftime("%d.%m.%Y %H:%M") if o["created_at"] else "",
                o["full_name"] or "",
                f"@{o['username']}" if o["username"] else "",
                o["user_phone"] or "",
            ])

        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf1 = io.BytesIO()
        wb1.save(buf1)
        buf1.seek(0)

        # ── Mahsulotlar ─────────────────────────────────────
        products = await get_all_products_for_export()
        wb2      = openpyxl.Workbook()
        ws2      = wb2.active
        ws2.title = "Mahsulotlar"

        headers2 = [
            "ID", "Nom", "Kategoriya", "Narx",
            "Stok", "Chegirma", "Faol", "Qo'shilgan"
        ]
        ws2.append(headers2)
        for cell in ws2[1]:
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center")

        for p in products:
            discount_text = ""
            if p["discount_active"] and p["discount_value"]:
                suffix = "%" if p["discount_type"] == "percent" else " so'm"
                discount_text = f"{p['discount_value']}{suffix}"

            ws2.append([
                p["id"],
                p["name"],
                p["category"] or "",
                float(p["price"]),
                p["stock_qty"],
                discount_text,
                "Ha" if p["is_active"] else "Yo'q",
                p["created_at"].strftime("%d.%m.%Y") if p["created_at"] else "",
            ])

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf2 = io.BytesIO()
        wb2.save(buf2)
        buf2.seek(0)

        # Yuborish
        await bot.send_message(
            chat_id=ADMIN_ID,
            text=f"💾 *Kunlik bekap* — {today}",
            parse_mode="Markdown"
        )
        await bot.send_document(
            chat_id=ADMIN_ID,
            document=BufferedInputFile(
                buf1.read(),
                filename=f"buyurtmalar_{today}.xlsx"
            ),
            caption="📋 Buyurtmalar"
        )
        await bot.send_document(
            chat_id=ADMIN_ID,
            document=BufferedInputFile(
                buf2.read(),
                filename=f"mahsulotlar_{today}.xlsx"
            ),
            caption="📦 Mahsulotlar"
        )
        logger.info(f"Kunlik bekap yuborildi: {today}")

    except Exception as e:
        logger.error(f"Bekap xatosi: {e}")
        try:
            await bot.send_message(
                chat_id=ADMIN_ID,
                text=f"❌ Bekap xatosi: {e}"
            )
        except Exception:
            pass


# ============================================================
# SCHEDULER — Ishga tushirish va to'xtatish
# ============================================================

def create_scheduler(bot: Bot) -> AsyncIOScheduler:
    """
    Barcha vazifalarni ro'yxatdan o'tkazib
    scheduler ni qaytaradi.
    main.py da ishga tushiriladi.
    """
    scheduler = AsyncIOScheduler(timezone="Asia/Tashkent")

    # Har 1 daqiqada — chegirmalar
    scheduler.add_job(
        check_expired_discounts,
        trigger=IntervalTrigger(minutes=1),
        id="check_discounts",
        name="Chegirmalar tekshirish",
        replace_existing=True
    )

    # Har kuni 09:00 — stok ogohlantirish
    scheduler.add_job(
        low_stock_alert,
        trigger=CronTrigger(hour=9, minute=0),
        args=[bot],
        id="low_stock_alert",
        name="Stok ogohlantirish",
        replace_existing=True
    )

    # Har kuni 03:00 — bekap
    scheduler.add_job(
        daily_backup,
        trigger=CronTrigger(hour=3, minute=0),
        args=[bot],
        id="daily_backup",
        name="Kunlik bekap",
        replace_existing=True
    )

    return scheduler
