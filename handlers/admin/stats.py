# ============================================================
# BOTFORGE — handlers/admin/stats.py
# Ma'lumotlar bo'limi — statistika, daromad, top mahsulotlar,
# Excel bekap, sozlamalar
# ============================================================

import logging
import io
from datetime import datetime
from aiogram import Router, F
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.types import InlineKeyboardButton
from aiogram.types import CallbackQuery, Message, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from db import (
    get_full_stats, get_revenue_stats,
    get_top_products, get_least_products,
    get_all_orders_for_export, get_all_products_for_export,
    get_setting, set_setting,
)
from keyboards import (
    AdminCB, StatsAdminCB,
    stats_menu_kb, stats_back_kb, back_to_main_kb,
)
from utils import fmt_price, fmt_date_short

logger = logging.getLogger(__name__)
router = Router()


# ============================================================
# FSM — Sozlamalar tahrirlash
# ============================================================

class SettingsState(StatesGroup):
    choose_field = State()
    new_value    = State()


SETTINGS_FIELDS = {
    "shop_name":    "🏪 Do'kon nomi",
    "shop_phone":   "📞 Telefon raqam",
    "shop_address": "📍 Manzil",
    "delivery_info":"🚚 Yetkazib berish",
    "payment_info": "💳 To'lov ma'lumoti",
    "payment_card": "💳 Karta raqami",
}


# ============================================================
# UMUMIY STATISTIKA
# ============================================================

@router.callback_query(StatsAdminCB.filter(F.section == "general"))
async def cb_stats_general(callback: CallbackQuery) -> None:
    stats = await get_full_stats()

    text = (
        f"📊 *Umumiy statistika*\n\n"
        f"👥 Foydalanuvchilar: *{stats['total_users']}* ta\n"
        f"🚫 Bloklangan: *{stats['banned_users']}* ta\n\n"
        f"📦 Mahsulotlar: *{stats['total_products']}* ta\n"
        f"✅ Faol: *{stats['active_products']}* ta\n"
        f"💰 Ombor qiymati: *{fmt_price(stats['total_stock_value'])}*\n\n"
        f"🛍 Jami buyurtmalar: *{stats['total_orders']}* ta\n"
        f"📦 Yetkazilgan: *{stats['delivered_orders']}* ta\n"
        f"❌ Bekor: *{stats['cancelled_orders']}* ta\n\n"
        f"💳 Jami daromad: *{fmt_price(stats['total_revenue'])}*"
    )

    await callback.message.edit_text(
        text=text,
        reply_markup=stats_back_kb(),
        parse_mode="Markdown"
    )
    await callback.answer()


# ============================================================
# DAROMAD
# ============================================================

@router.callback_query(StatsAdminCB.filter(F.section == "revenue"))
async def cb_stats_revenue(callback: CallbackQuery) -> None:
    revenue = await get_revenue_stats()

    text = (
        f"💰 *Daromad*\n\n"
        f"📅 Bugun: *{fmt_price(revenue['today'])}*\n"
        f"📅 Bu hafta: *{fmt_price(revenue['this_week'])}*\n"
        f"📅 Bu oy: *{fmt_price(revenue['this_month'])}*\n"
        f"📅 Jami: *{fmt_price(revenue['total'])}*"
    )

    await callback.message.edit_text(
        text=text,
        reply_markup=stats_back_kb(),
        parse_mode="Markdown"
    )
    await callback.answer()


# ============================================================
# TOP MAHSULOTLAR
# ============================================================

@router.callback_query(StatsAdminCB.filter(F.section == "top"))
async def cb_stats_top(callback: CallbackQuery) -> None:
    products = await get_top_products(limit=10)

    if not products:
        await callback.message.edit_text(
            text="🏆 Hali sotilgan mahsulotlar yo'q.",
            reply_markup=stats_back_kb()
        )
        await callback.answer()
        return

    lines = ["🏆 *Eng ko'p sotilgan mahsulotlar:*\n"]
    for i, p in enumerate(products, 1):
        lines.append(
            f"{i}. *{p['name']}*\n"
            f"   📦 {p['total_qty']} dona | 💰 {fmt_price(p['total_revenue'])}"
        )

    await callback.message.edit_text(
        text="\n".join(lines),
        reply_markup=stats_back_kb(),
        parse_mode="Markdown"
    )
    await callback.answer()


# ============================================================
# KAM SOTILGAN MAHSULOTLAR
# ============================================================

@router.callback_query(StatsAdminCB.filter(F.section == "least"))
async def cb_stats_least(callback: CallbackQuery) -> None:
    products = await get_least_products(limit=10)

    if not products:
        await callback.message.edit_text(
            text="📉 Ma'lumot yo'q.",
            reply_markup=stats_back_kb()
        )
        await callback.answer()
        return

    lines = ["📉 *Kam sotilgan mahsulotlar:*\n"]
    for p in products:
        lines.append(
            f"• *{p['name']}*\n"
            f"   📦 Stok: {p['stock_qty']} | Sotilgan: {p['total_qty']} dona"
        )

    await callback.message.edit_text(
        text="\n".join(lines),
        reply_markup=stats_back_kb(),
        parse_mode="Markdown"
    )
    await callback.answer()


# ============================================================
# EXCEL BEKAP
# ============================================================

@router.callback_query(StatsAdminCB.filter(F.section == "export"))
async def cb_export(callback: CallbackQuery) -> None:
    """
    Ikki Excel fayl yaratib yuboradi:
    1. Buyurtmalar
    2. Mahsulotlar
    """
    await callback.answer("⏳ Excel tayyorlanmoqda...")
    await callback.message.edit_text(
        text="⏳ Excel fayllar tayyorlanmoqda...",
        reply_markup=None
    )

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        today = fmt_date_short(datetime.now())

        # ── 1. Buyurtmalar ──────────────────────────────────
        orders = await get_all_orders_for_export()
        wb_orders = openpyxl.Workbook()
        ws = wb_orders.active
        ws.title = "Buyurtmalar"

        headers = [
            "Kod", "Holat", "Summa", "Manzil",
            "Telefon", "Izoh", "Sana",
            "Ism", "Username", "Foydalanuvchi tel"
        ]
        ws.append(headers)

        # Sarlavha uslubi
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E86AB")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for o in orders:
            ws.append([
                o["order_code"],
                o["status"],
                float(o["total_price"] or 0),
                o["address"] or "",
                o["order_phone"] or "",
                o["note"] or "",
                o["created_at"].strftime("%d.%m.%Y %H:%M") if o["created_at"] else "",
                o["full_name"] or "",
                f"@{o['username']}" if o["username"] else "",
                o["user_phone"] or "",
            ])

        # Ustun kengligini avtomatik sozlash
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        orders_buf = io.BytesIO()
        wb_orders.save(orders_buf)
        orders_buf.seek(0)

        # ── 2. Mahsulotlar ──────────────────────────────────
        products = await get_all_products_for_export()
        wb_prods = openpyxl.Workbook()
        ws2 = wb_prods.active
        ws2.title = "Mahsulotlar"

        headers2 = [
            "ID", "Nom", "Kategoriya", "Narx",
            "Stok", "Chegirma", "Faol", "Qo'shilgan"
        ]
        ws2.append(headers2)

        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for p in products:
            discount_text = ""
            if p["discount_active"] and p["discount_value"]:
                suffix = "%" if p["discount_type"] == "percent" else " so'm"
                discount_text = f"{p['discount_value']}{suffix}"

            ws2.append([
                p["id"],
                p["name"],
                p["category"] or "",
                float(p["price"]),
                p["stock_qty"],
                discount_text,
                "Ha" if p["is_active"] else "Yo'q",
                p["created_at"].strftime("%d.%m.%Y") if p["created_at"] else "",
            ])

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        prods_buf = io.BytesIO()
        wb_prods.save(prods_buf)
        prods_buf.seek(0)

        # Fayllarni yuborish
        await callback.message.answer_document(
            document=BufferedInputFile(
                orders_buf.read(),
                filename=f"buyurtmalar_{today}.xlsx"
            ),
            caption=f"📋 Buyurtmalar — {today}"
        )
        await callback.message.answer_document(
            document=BufferedInputFile(
                prods_buf.read(),
                filename=f"mahsulotlar_{today}.xlsx"
            ),
            caption=f"📦 Mahsulotlar — {today}"
        )

        await callback.message.edit_text(
            text="✅ Excel fayllar yuborildi!",
            reply_markup=stats_back_kb()
        )

    except Exception as e:
        logger.error(f"Excel export xatosi: {e}")
        await callback.message.edit_text(
            text="❌ Excel yaratishda xatolik yuz berdi.",
            reply_markup=stats_back_kb()
        )


# ============================================================
# SOZLAMALAR
# ============================================================

@router.callback_query(StatsAdminCB.filter(F.section == "settings"))
async def cb_settings(
    callback: CallbackQuery,
    state: FSMContext
) -> None:
    """Do'kon sozlamalari ro'yxati."""
    await state.clear()
    lines = ["⚙️ *Sozlamalar*\n"]

    for key, label in SETTINGS_FIELDS.items():
        value = await get_setting(key)
        lines.append(f"{label}: `{value or '—'}`")

    builder = InlineKeyboardBuilder()
    for key, label in SETTINGS_FIELDS.items():
        builder.row(InlineKeyboardButton(
            text=f"✏️ {label}",
            callback_data=f"settings_edit:{key}"
        ))
    builder.row(
        InlineKeyboardButton(
            text="🔙 Orqaga",
            callback_data=AdminCB(section="stats").pack()
        ),
        InlineKeyboardButton(
            text="🏠 Bosh menyu",
            callback_data=AdminCB(section="main").pack()
        )
    )

    await callback.message.edit_text(
        text="\n".join(lines),
        reply_markup=builder.as_markup(),
        parse_mode="Markdown"
    )
    await callback.answer()


@router.callback_query(F.data.startswith("settings_edit:"))
async def cb_settings_edit(
    callback: CallbackQuery,
    state: FSMContext
) -> None:
    key = callback.data.split(":")[1]
    label = SETTINGS_FIELDS.get(key, key)
    current = await get_setting(key)

    await state.update_data(settings_key=key)
    await state.set_state(SettingsState.new_value)

    await callback.message.edit_text(
        text=(
            f"✏️ *{label}*\n\n"
            f"Hozirgi qiymat: `{current or '—'}`\n\n"
            f"Yangi qiymatni yozing:"
        ),
        reply_markup=None,
        parse_mode="Markdown"
    )
    await callback.answer()


@router.message(SettingsState.new_value, F.text)
async def process_settings_value(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    await state.clear()

    key   = data["settings_key"]
    value = message.text.strip()
    label = SETTINGS_FIELDS.get(key, key)

    await set_setting(key, value)
    await message.answer(
        f"✅ *{label}* yangilandi!\n\n`{value}`",
        parse_mode="Markdown"
    )
